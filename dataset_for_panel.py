import pandas as pd
import numpy as np
import openpyxl

TXNS_FILE = "transactions_v7_good.csv"
VIX_FILE = "VIXCLS_daily.xlsx"
DGS10_FILE = "DGS10_daily.xlsx"
DGS1MO_FILE = "DGS1MO_daily.xlsx"
ADS_FILE = "ADS_Index_Most_Current_Vintage.xlsx"
ETH_FILE = "eth-usd-max.csv"
OUTPUT_FILE = "transaction_panel_v4.xlsx"

tx = pd.read_csv(TXNS_FILE)

# Get date from round_time
tx['round_time'] = pd.to_datetime(tx['round_time'], format='mixed', dayfirst=False)
tx['date'] = tx['round_time'].dt.normalize()

# Keep only transactions where real_transfer_amount > 0
tx = tx[tx['real_transfer_amount'] > 0].copy()

# Exclude known blockchain anomaly - LFTY0315 on 2023-08-21
# real_transfer_amount = 999,995,000,000 — misclassified asset configuration event, not a real transfer
ANOMALOUS_TX_IDS = {
    'XEUUH5WVQDV5ODHBX5BATBHFLLNANJDKESBRGBPXK6WCYGEOHWOA',
}
tx = tx[~tx['tx_id'].isin(ANOMALOUS_TX_IDS)].copy()

# Sum real_transfer_amount per asa_id per trading day
# Each asa_id (property) starts from its own first trading day to its last
daily = (
    tx.groupby(['asa_id', 'date'], as_index=False)
    .agg(total_tokens_transferred=('real_transfer_amount', 'sum'),
         n_transactions=('tx_id', 'nunique'))
)

daily = daily.sort_values(['asa_id', 'date']).reset_index(drop=True)


# To load data from FRED
def load_fred_xlsx(filepath, sheet_name, col_name):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]
    data = [(r[0], r[1]) for r in ws.iter_rows(min_row=2, values_only=True)]
    df = pd.DataFrame(data, columns=['date', col_name])
    df['date'] = pd.to_datetime(df['date'])
    df[col_name] = pd.to_numeric(df[col_name], errors='coerce')
    return df.sort_values('date').reset_index(drop=True)


vix = load_fred_xlsx(VIX_FILE,   'Daily, Close', 'vix')
dgs10 = load_fred_xlsx(DGS10_FILE, 'Daily', 'dgs10')
dgs1mo = load_fred_xlsx(DGS1MO_FILE, 'Daily', 'dgs1mo')


def build_delta_series(df, col):
    """
      - For Saturday: delta = Friday - Thursday
      - For Sunday: delta = Friday - Thursday
      - For Monday: delta = Monday - Friday
    """
    # Set date as index
    s = df.set_index('date')[col].copy()

    # Drop NaN rows
    s_trading = s.dropna().sort_index()

    # Compute delta between consecutive trading days
    delta_trading = s_trading.diff()

    # Reindex to full calendar range
    full_idx = pd.date_range(s_trading.index.min(), s_trading.index.max(), freq='D')
    delta_calendar = delta_trading.reindex(full_idx)

    # Weekends and holidays carry the most recent trading day's delta
    # This means Saturday & Sunday carries Friday's delta, Monday gets its own delta (Monday - Friday)
    delta_filled = delta_calendar.ffill()

    result = pd.DataFrame({'date': full_idx, f'delta_{col}': delta_filled.values})
    return result


delta_vix = build_delta_series(vix,   'vix')
delta_dgs10 = build_delta_series(dgs10, 'dgs10')
delta_dgs1mo = build_delta_series(dgs1mo, 'dgs1mo')

# Load and add ADS data based on date
try:
    wb_ads = openpyxl.load_workbook(ADS_FILE)
    ws_ads = wb_ads.active
    ads_data = [(r[0], r[1]) for r in ws_ads.iter_rows(min_row=2, values_only=True)
                if r[0] is not None and r[1] is not None]
    ads = pd.DataFrame(ads_data, columns=['date', 'ads'])

    # ADS date formatting
    ads['date'] = pd.to_datetime(ads['date'], format='%Y:%m:%d')
    ads['ads'] = pd.to_numeric(ads['ads'], errors='coerce')
    ads = ads.sort_values('date').reset_index(drop=True)

    full_idx_ads = pd.date_range(ads['date'].min(), ads['date'].max(), freq='D')
    ads = ads.set_index('date').reindex(full_idx_ads).ffill().reset_index()
    ads.columns = ['date', 'ads']
except Exception as e:
    print(f"  ADS error: {e}")
    ads = None

# Load ETH prices data
eth = pd.read_csv(ETH_FILE)
eth['date'] = pd.to_datetime(eth['snapped_at'].str.replace(' UTC', ''), format='mixed')
eth['date'] = eth['date'].dt.normalize()
eth['price'] = pd.to_numeric(eth['price'], errors='coerce')
eth = eth[['date', 'price']].sort_values('date').drop_duplicates('date').reset_index(drop=True)

# Calculate simple daily return: (Pt / Pt-1) - 1
eth['eth_return'] = (eth['price'] / eth['price'].shift(1)) - 1
eth_full = eth[['date', 'eth_return']].copy()

# Merge the datas
panel = daily.copy()
panel = panel.merge(delta_vix,    on='date', how='left')
panel = panel.merge(delta_dgs10,  on='date', how='left')
panel = panel.merge(delta_dgs1mo, on='date', how='left')

if ads is not None:
    panel = panel.merge(ads[['date', 'ads']], on='date', how='left')
else:
    panel['ads'] = np.nan

panel = panel.merge(eth_full[['date', 'eth_return']], on='date', how='left')

panel = panel[[
    'asa_id',
    'date',
    'total_tokens_transferred',
    'n_transactions',
    'delta_vix',
    'delta_dgs10',
    'delta_dgs1mo',
    'ads',
    'eth_return',
]]

panel = panel.sort_values(['asa_id', 'date']).reset_index(drop=True)

# Save
panel.to_excel(OUTPUT_FILE, index=False)
print(f"Saved {OUTPUT_FILE}")